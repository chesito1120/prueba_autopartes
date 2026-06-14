from flask import Flask, render_template, request, redirect, session
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from urllib.parse import quote_plus, urlencode
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os
import uuid
import zipfile
from datetime import datetime, timedelta
import requests

app = Flask(__name__)

app.secret_key = os.getenv("SECRET_KEY", "autopartes_secret")

# =========================
# MERCADO LIBRE
# =========================

MELI_APP_ID = os.getenv("MELI_APP_ID", "516751596401763")
MELI_CLIENT_SECRET = os.getenv("MELI_CLIENT_SECRET", "aLvmsfXSpLcavxUrYFfwPLCr9y4Kio5X")
MELI_REDIRECT_URI = os.getenv(
    "MELI_REDIRECT_URI",
    "https://prueba-autopartes.onrender.com/mercadolibre/callback"
)

MELI_AUTH_URL = "https://auth.mercadolibre.com.mx/authorization"
MELI_TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
MELI_API_URL = "https://api.mercadolibre.com"

CRON_SECRET = os.getenv("CRON_SECRET", "autopartes_ch_sync_2026")

# =========================
# BASE DE DATOS
# =========================

MYSQL_URL = os.getenv("MYSQL_URL") or os.getenv("DATABASE_URL")

DB_USER = os.getenv("MYSQLUSER")
DB_PASSWORD = os.getenv("MYSQLPASSWORD")
DB_HOST = os.getenv("MYSQLHOST")
DB_PORT = os.getenv("MYSQLPORT")
DB_NAME = os.getenv("MYSQLDATABASE")

print("========== DB DEBUG ==========")
print("MYSQL_URL existe:", bool(MYSQL_URL))
print("MYSQLUSER existe:", bool(DB_USER))
print("MYSQLPASSWORD existe:", bool(DB_PASSWORD))
print("MYSQLHOST existe:", bool(DB_HOST))
print("MYSQLPORT existe:", bool(DB_PORT))
print("MYSQLDATABASE existe:", bool(DB_NAME))
print("==============================")

if MYSQL_URL:
    app.config["SQLALCHEMY_DATABASE_URI"] = MYSQL_URL

elif all([DB_USER, DB_PASSWORD, DB_HOST, DB_PORT, DB_NAME]):
    DB_PASSWORD_SAFE = quote_plus(DB_PASSWORD)

    app.config["SQLALCHEMY_DATABASE_URI"] = (
        f"mysql+pymysql://{DB_USER}:{DB_PASSWORD_SAFE}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    )

else:
    print("ADVERTENCIA: No se encontraron variables MySQL completas. Usando SQLite temporal.")
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///autopartes_local.db"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = "static/uploads"

db = SQLAlchemy(app)

print("BASE DE DATOS ACTIVA:")
print(app.config["SQLALCHEMY_DATABASE_URI"])


# =========================
# FUNCIONES DE IMAGEN
# =========================

def guardar_imagen(archivo):
    if not archivo or not archivo.filename:
        return ""

    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    nombre_original = secure_filename(archivo.filename)
    extension = nombre_original.rsplit(".", 1)[-1].lower()

    if extension not in ["jpg", "jpeg", "png", "webp"]:
        return ""

    nombre_archivo = f"{uuid.uuid4().hex}.{extension}"
    ruta = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)

    archivo.save(ruta)

    return nombre_archivo


def guardar_varias_imagenes(archivos):
    fotos = []

    for archivo in archivos:
        if archivo and archivo.filename:
            nombre = guardar_imagen(archivo)

            if nombre:
                fotos.append(nombre)

        if len(fotos) == 5:
            break

    return "|".join(fotos)


def descargar_imagen_ml(url_imagen):
    if not url_imagen:
        return ""

    try:
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

        respuesta = requests.get(url_imagen, timeout=20)

        if respuesta.status_code != 200:
            return ""

        content_type = respuesta.headers.get("Content-Type", "").lower()

        if "jpeg" in content_type or "jpg" in content_type:
            extension = "jpg"
        elif "png" in content_type:
            extension = "png"
        elif "webp" in content_type:
            extension = "webp"
        else:
            extension = "jpg"

        nombre_archivo = f"ml_{uuid.uuid4().hex}.{extension}"
        ruta = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)

        with open(ruta, "wb") as archivo:
            archivo.write(respuesta.content)

        return nombre_archivo

    except Exception as e:
        print("ERROR descargando imagen ML:", e)
        return ""


def descargar_imagenes_ml(item):
    fotos = []

    pictures = item.get("pictures") or []

    for picture in pictures:
        url_imagen = (
            picture.get("secure_url")
            or picture.get("url")
            or picture.get("max_size")
        )

        nombre = descargar_imagen_ml(url_imagen)

        if nombre:
            fotos.append(nombre)

        if len(fotos) == 5:
            break

    if not fotos:
        thumbnail = item.get("secure_thumbnail") or item.get("thumbnail")

        nombre = descargar_imagen_ml(thumbnail)

        if nombre:
            fotos.append(nombre)

    return "|".join(fotos)


def extraer_zip_imagenes(archivo_zip):
    imagenes_extraidas = set()

    if not archivo_zip or not archivo_zip.filename:
        return imagenes_extraidas

    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    nombre_zip = secure_filename(archivo_zip.filename)
    ruta_zip = os.path.join(app.config["UPLOAD_FOLDER"], nombre_zip)

    archivo_zip.save(ruta_zip)

    try:
        with zipfile.ZipFile(ruta_zip, "r") as zip_ref:
            for archivo in zip_ref.namelist():
                nombre_archivo = os.path.basename(archivo)

                if not nombre_archivo:
                    continue

                nombre_seguro = secure_filename(nombre_archivo)

                if "." not in nombre_seguro:
                    continue

                extension = nombre_seguro.rsplit(".", 1)[-1].lower()

                if extension not in ["jpg", "jpeg", "png", "webp"]:
                    continue

                ruta_destino = os.path.join(app.config["UPLOAD_FOLDER"], nombre_seguro)

                with zip_ref.open(archivo) as origen, open(ruta_destino, "wb") as destino:
                    destino.write(origen.read())

                imagenes_extraidas.add(nombre_seguro)

    except zipfile.BadZipFile:
        print("ERROR: El archivo subido no es un ZIP válido.")

    if os.path.exists(ruta_zip):
        os.remove(ruta_zip)

    return imagenes_extraidas


# =========================
# FUNCIONES DE CRÉDITO Y MOVIMIENTOS
# =========================

def calcular_abonado_credito(credito_id):
    total_abonado = db.session.query(
        db.func.coalesce(db.func.sum(AbonoCreditoVenta.monto), 0)
    ).filter(
        AbonoCreditoVenta.credito_id == credito_id
    ).scalar()

    return round(float(total_abonado or 0), 2)


def calcular_saldo_credito(credito):
    abonado = calcular_abonado_credito(credito.id)
    total = credito.total or 0
    saldo = round(max(total - abonado, 0), 2)
    return abonado, saldo


def calcular_total_movimiento(producto, cantidad):
    precio = producto.costo_venta or 0
    return round(precio * cantidad, 2)


# =========================
# MODELOS
# =========================

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    nombre = db.Column(db.String(150), nullable=False)
    correo = db.Column(db.String(180), unique=True, nullable=False)
    password_hash = db.Column(db.String(300), nullable=False)
    rol = db.Column(db.String(50), default="vendedor")
    activo = db.Column(db.Boolean, default=True)


class MercadoLibreToken(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    user_id = db.Column(db.String(100))
    access_token = db.Column(db.Text)
    refresh_token = db.Column(db.Text)
    expires_at = db.Column(db.DateTime)

    scope = db.Column(db.String(300))
    token_type = db.Column(db.String(50))

    fecha_conexion = db.Column(db.String(100))
    fecha_actualizacion = db.Column(db.String(100))


class MercadoLibreProducto(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    ml_item_id = db.Column(db.String(100), unique=True, nullable=False)

    producto_id = db.Column(db.Integer, db.ForeignKey("producto.id"), nullable=False)
    producto = db.relationship("Producto", backref="mercadolibre_sync")

    ml_status = db.Column(db.String(100))
    ml_permalink = db.Column(db.String(500))
    ml_last_updated = db.Column(db.String(100))
    fecha_sync = db.Column(db.String(100))


class Movimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    producto_id = db.Column(db.Integer, db.ForeignKey("producto.id"), nullable=False)
    producto = db.relationship("Producto", backref="movimientos")

    tipo = db.Column(db.String(50))
    estado = db.Column(db.String(50))

    cliente = db.Column(db.String(150))
    cantidad = db.Column(db.Integer, default=1)
    precio_unitario = db.Column(db.Float, default=0)
    total = db.Column(db.Float, default=0)

    metodo_pago = db.Column(db.String(100))
    factura = db.Column(db.String(100))
    fecha_salida = db.Column(db.String(100))
    fecha_prestamo = db.Column(db.String(100))
    comentarios = db.Column(db.Text)
    fecha_registro = db.Column(db.String(100))


class CreditoVenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    movimiento_id = db.Column(db.Integer, db.ForeignKey("movimiento.id"), nullable=False)
    movimiento = db.relationship("Movimiento", backref="credito")

    producto_id = db.Column(db.Integer, db.ForeignKey("producto.id"), nullable=False)
    producto = db.relationship("Producto", backref="creditos_venta")

    cliente = db.Column(db.String(150))
    cantidad = db.Column(db.Integer, default=1)
    precio_unitario = db.Column(db.Float, default=0)
    total = db.Column(db.Float, default=0)
    estado = db.Column(db.String(50), default="credito")
    fecha_salida = db.Column(db.String(100))
    comentarios = db.Column(db.Text)


class AbonoCreditoVenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    credito_id = db.Column(db.Integer, db.ForeignKey("credito_venta.id"), nullable=False)
    credito = db.relationship("CreditoVenta", backref="abonos")

    monto = db.Column(db.Float, default=0)
    fecha = db.Column(db.String(100))
    comentario = db.Column(db.String(300))


class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    marca = db.Column(db.String(100))
    modelo = db.Column(db.String(100))
    anio = db.Column(db.Integer)
    autoparte = db.Column(db.String(150))
    observaciones = db.Column(db.Text)
    transmision = db.Column(db.String(100))
    motor = db.Column(db.String(100))
    propiedad = db.Column(db.String(100))
    tipo = db.Column(db.String(100))
    lado = db.Column(db.String(50))
    origen = db.Column(db.String(100))
    costo_venta = db.Column(db.Float)
    stock = db.Column(db.Integer, default=1)
    foto = db.Column(db.String(300))
    link_ml = db.Column(db.String(500))
    estado = db.Column(db.String(50), default="disponible")

    vendido_a = db.Column(db.String(150))
    metodo_pago = db.Column(db.String(100))
    factura = db.Column(db.String(100))
    comentarios_venta = db.Column(db.Text)
    fecha_salida = db.Column(db.String(100))
    fecha_prestamo = db.Column(db.String(100))


# =========================
# FUNCIONES MERCADO LIBRE
# =========================

def obtener_token_ml():
    token = MercadoLibreToken.query.order_by(MercadoLibreToken.id.desc()).first()

    if not token:
        return None, "No existe conexión con Mercado Libre."

    if not token.access_token:
        return None, "No existe access token de Mercado Libre."

    if token.expires_at and datetime.now() < token.expires_at - timedelta(minutes=10):
        return token, None

    if not token.refresh_token:
        return None, "No existe refresh token para renovar Mercado Libre."

    payload = {
        "grant_type": "refresh_token",
        "client_id": MELI_APP_ID,
        "client_secret": MELI_CLIENT_SECRET,
        "refresh_token": token.refresh_token
    }

    headers = {
        "accept": "application/json",
        "content-type": "application/x-www-form-urlencoded"
    }

    respuesta = requests.post(
        MELI_TOKEN_URL,
        data=payload,
        headers=headers,
        timeout=20
    )

    if respuesta.status_code not in [200, 201]:
        print("ERROR REFRESH TOKEN ML:")
        print(respuesta.status_code)
        print(respuesta.text)
        return None, f"Error renovando token ML: {respuesta.text}"

    datos = respuesta.json()

    token.access_token = datos.get("access_token")
    token.refresh_token = datos.get("refresh_token")
    token.expires_at = datetime.now() + timedelta(seconds=datos.get("expires_in", 0))
    token.scope = datos.get("scope")
    token.token_type = datos.get("token_type")
    token.fecha_actualizacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    db.session.commit()

    return token, None


def ml_get(url, token):
    headers = {
        "Authorization": f"Bearer {token.access_token}"
    }

    return requests.get(
        url,
        headers=headers,
        timeout=30
    )


def obtener_ids_publicaciones_ml(token):
    item_ids = []
    offset = 0
    limit = 50

    while True:
        url = (
            f"{MELI_API_URL}/users/{token.user_id}/items/search"
            f"?status=active"
            f"&limit={limit}"
            f"&offset={offset}"
        )

        respuesta = ml_get(url, token)

        if respuesta.status_code != 200:
            return item_ids, f"Error obteniendo publicaciones: {respuesta.text}"

        datos = respuesta.json()
        resultados = datos.get("results", [])

        if not resultados:
            break

        item_ids.extend(resultados)

        paging = datos.get("paging", {})
        total = paging.get("total", 0)

        offset += limit

        if offset >= total:
            break

        if len(item_ids) >= 1000:
            break

    return item_ids, None


def obtener_detalles_items_ml(token, item_ids):
    detalles = []

    for i in range(0, len(item_ids), 20):
        bloque = item_ids[i:i + 20]
        ids = ",".join(bloque)

        attributes = ",".join([
            "id",
            "title",
            "price",
            "available_quantity",
            "permalink",
            "thumbnail",
            "secure_thumbnail",
            "pictures",
            "status",
            "last_updated",
            "category_id",
            "attributes"
        ])

        url = f"{MELI_API_URL}/items?ids={ids}&attributes={attributes}"

        respuesta = ml_get(url, token)

        if respuesta.status_code != 200:
            print("ERROR multiget ML:", respuesta.text)
            continue

        datos = respuesta.json()

        for item_respuesta in datos:
            if item_respuesta.get("code") == 200:
                body = item_respuesta.get("body") or {}
                detalles.append(body)

    return detalles


def obtener_atributo_ml(item, atributo_id):
    atributos = item.get("attributes") or []

    for atributo in atributos:
        if atributo.get("id") == atributo_id:
            return atributo.get("value_name") or ""

    return ""

def obtener_atributo_variacion_ml(item, atributo_id):
    variaciones = item.get("variations") or []

    for variacion in variaciones:
        combinaciones = variacion.get("attribute_combinations") or []

        for atributo in combinaciones:
            if atributo.get("id") == atributo_id:
                return atributo.get("value_name") or ""

    return ""


def sincronizar_publicaciones_mercadolibre():
    token, error = obtener_token_ml()

    if error:
        return error

    item_ids, error_ids = obtener_ids_publicaciones_ml(token)

    if error_ids:
        return error_ids

    if not item_ids:
        return "Sincronización finalizada. No se encontraron publicaciones activas en Mercado Libre."

    detalles = obtener_detalles_items_ml(token, item_ids)

    creados = 0
    actualizados = 0
    pausados = 0
    errores = 0

    ids_activos_ml = set()

    for item in detalles:
        try:
            ml_item_id = item.get("id")

            if not ml_item_id:
                errores += 1
                continue

            ids_activos_ml.add(ml_item_id)

            status = item.get("status") or ""
            titulo = item.get("title") or "Publicación Mercado Libre"
            precio = float(item.get("price") or 0)
            stock = int(item.get("available_quantity") or 0)
            permalink = item.get("permalink") or ""
            last_updated = item.get("last_updated") or ""

            marca = (
                obtener_atributo_ml(item, "BRAND")
                or obtener_atributo_ml(item, "MANUFACTURER"))

            modelo = (
                obtener_atributo_ml(item, "MODEL")
                or obtener_atributo_ml(item, "VEHICLE_MODEL")
                or obtener_atributo_ml(item, "CAR_MODEL")
            )

            anio = (
                obtener_atributo_ml(item, "YEAR")
                or obtener_atributo_ml(item, "VEHICLE_YEAR")
            )

            lado = (
                obtener_atributo_ml(item, "SIDE")
                or obtener_atributo_ml(item, "POSITION")
                or obtener_atributo_ml(item, "VEHICLE_PARTS_POSITION")
                or obtener_atributo_variacion_ml(item, "VEHICLE_PARTS_POSITION")
            )
            tipo = (
                obtener_atributo_ml(item, "VEHICLE_TYPE")
                or obtener_atributo_ml(item, "ITEM_CONDITION")
            )

            numero_parte = obtener_atributo_ml(item, "PART_NUMBER")

            motor = (
                obtener_atributo_ml(item, "ENGINE")
                or obtener_atributo_ml(item, "MOTOR")
            )

            transmision = (
                obtener_atributo_ml(item, "TRANSMISSION")
            )

            sync = MercadoLibreProducto.query.filter_by(
                ml_item_id=ml_item_id
            ).first()

            producto = None

            if sync:
                producto = Producto.query.get(sync.producto_id)

                if producto:
                    actualizados += 1
                else:
                    producto = Producto()
                    db.session.add(producto)
                    db.session.flush()

                    sync.producto_id = producto.id
                    actualizados += 1
            else:
                producto = Producto()
                db.session.add(producto)
                db.session.flush()

                sync = MercadoLibreProducto(
                    ml_item_id=ml_item_id,
                    producto_id=producto.id
                )

                db.session.add(sync)
                creados += 1

            producto.autoparte = titulo
            producto.costo_venta = precio
            producto.stock = stock
            producto.link_ml = permalink
            producto.estado = "disponible" if stock > 0 and status == "active" else "agotado"

            if marca:
                producto.marca = marca

            if modelo:
                producto.modelo = modelo
            
            if anio:
                try:
                    producto.anio = int(str(anio).split("-")[0].strip())
                except:
                    pass

            if lado:
                producto.lado = lado

            if motor:
                producto.motor = motor

            if transmision:
                producto.transmision = transmision

            producto.origen = "Mercado Libre"
            producto.tipo = producto.tipo or "Mercado Libre"
            producto.propiedad = producto.propiedad or "SIN PROPIEDAD"

            if tipo:
                producto.tipo = tipo

            if numero_parte:
                producto.observaciones = f"Número de parte: {numero_parte}"

            if not producto.foto:
                fotos_ml = descargar_imagenes_ml(item)

                if fotos_ml:
                    producto.foto = fotos_ml

            sync.ml_status = status
            sync.ml_permalink = permalink
            sync.ml_last_updated = last_updated
            sync.fecha_sync = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        except Exception as e:
            print("ERROR sincronizando item ML:", e)
            errores += 1

    syncs = MercadoLibreProducto.query.all()

    for sync in syncs:
        producto = Producto.query.get(sync.producto_id)

        if sync.ml_item_id not in ids_activos_ml:
            if producto:
                producto.stock = 0
                producto.estado = "agotado"

            sync.ml_status = "no_activo"
            sync.fecha_sync = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            pausados += 1

    db.session.commit()

    return (
        f"Sincronización Mercado Libre finalizada. "
        f"Creados: {creados}. "
        f"Actualizados: {actualizados}. "
        f"Marcados como agotados/no activos: {pausados}. "
        f"Errores: {errores}. "
        f"Total publicaciones activas ML: {len(item_ids)}."
    )


# =========================
# CATÁLOGO
# =========================

@app.route("/")
def home():
    pagina = request.args.get("page", 1, type=int)
    por_pagina = 21

    productos_paginados = Producto.query.filter(
        Producto.stock > 0
    ).paginate(
        page=pagina,
        per_page=por_pagina,
        error_out=False
    )

    return render_template(
        "catalogo.html",
        productos=productos_paginados.items,
        paginacion=productos_paginados
    )


# =========================
# RUTAS MERCADO LIBRE
# =========================

@app.route("/mercadolibre")
def mercadolibre():
    if session.get("rol") != "admin":
        return redirect("/inventario")

    token = MercadoLibreToken.query.order_by(MercadoLibreToken.id.desc()).first()

    return render_template(
        "mercadolibre.html",
        token=token,
        meli_app_id=MELI_APP_ID,
        meli_redirect_uri=MELI_REDIRECT_URI
    )


@app.route("/mercadolibre/webhook", methods=["POST", "GET"])
def mercadolibre_webhook():
    if request.method == "GET":
        return "OK", 200

    data = request.get_json(silent=True) or {}

    print("WEBHOOK ML RECIBIDO:")
    print(data)

    return "OK", 200


@app.route("/mercadolibre/conectar")
def mercadolibre_conectar():
    if session.get("rol") != "admin":
        return redirect("/inventario")

    if not MELI_APP_ID or not MELI_REDIRECT_URI:
        return "Faltan variables MELI_APP_ID o MELI_REDIRECT_URI", 500

    params = {
        "response_type": "code",
        "client_id": MELI_APP_ID,
        "redirect_uri": MELI_REDIRECT_URI
    }

    url_autorizacion = f"{MELI_AUTH_URL}?{urlencode(params)}"

    return redirect(url_autorizacion)


@app.route("/mercadolibre/callback")
def mercadolibre_callback():
    if session.get("rol") != "admin":
        return redirect("/inventario")

    code = request.args.get("code")

    if not code:
        return "No se recibió code de Mercado Libre.", 400

    if not MELI_APP_ID or not MELI_CLIENT_SECRET or not MELI_REDIRECT_URI:
        return "Faltan variables de Mercado Libre en el servidor.", 500

    payload = {
        "grant_type": "authorization_code",
        "client_id": MELI_APP_ID,
        "client_secret": MELI_CLIENT_SECRET,
        "code": code,
        "redirect_uri": MELI_REDIRECT_URI
    }

    headers = {
        "accept": "application/json",
        "content-type": "application/x-www-form-urlencoded"
    }

    respuesta = requests.post(
        MELI_TOKEN_URL,
        data=payload,
        headers=headers,
        timeout=20
    )

    if respuesta.status_code not in [200, 201]:
        print("ERROR MERCADO LIBRE TOKEN:")
        print(respuesta.status_code)
        print(respuesta.text)
        return f"Error al obtener token de Mercado Libre: {respuesta.text}", 400

    datos = respuesta.json()

    expires_in = datos.get("expires_in", 0)
    expires_at = datetime.now() + timedelta(seconds=expires_in)

    token = MercadoLibreToken.query.order_by(MercadoLibreToken.id.desc()).first()

    if not token:
        token = MercadoLibreToken()

    token.user_id = str(datos.get("user_id", ""))
    token.access_token = datos.get("access_token")
    token.refresh_token = datos.get("refresh_token")
    token.expires_at = expires_at
    token.scope = datos.get("scope")
    token.token_type = datos.get("token_type")
    token.fecha_conexion = token.fecha_conexion or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    token.fecha_actualizacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    db.session.add(token)
    db.session.commit()

    return redirect("/mercadolibre")


@app.route("/mercadolibre/publicaciones")
def mercadolibre_publicaciones():
    token, error = obtener_token_ml()

    if error:
        return error

    respuesta = ml_get(
        f"{MELI_API_URL}/users/{token.user_id}/items/search?include_filters=true",
        token
    )

    return respuesta.text

@app.route("/mercadolibre/debug-item/<item_id>")
def mercadolibre_debug_item(item_id):
    token, error = obtener_token_ml()

    if error:
        return error

    respuesta = ml_get(
        f"{MELI_API_URL}/items/{item_id}",
        token
    )

    return respuesta.text


@app.route("/mercadolibre/sincronizar")
def mercadolibre_sincronizar():
    if session.get("rol") != "admin":
        return redirect("/inventario")

    resultado = sincronizar_publicaciones_mercadolibre()

    return resultado


@app.route("/mercadolibre/cron-sync")
def mercadolibre_cron_sync():
    secret = request.args.get("secret")

    if secret != CRON_SECRET:
        return "No autorizado", 401

    resultado = sincronizar_publicaciones_mercadolibre()

    return resultado


# =========================
# LOGIN
# =========================

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        correo = (request.form.get("correo") or "").strip().lower()
        password = request.form.get("password") or ""

        usuario = Usuario.query.filter_by(correo=correo).first()

        if usuario and usuario.activo and check_password_hash(usuario.password_hash, password):
            session["usuario"] = usuario.nombre
            session["correo"] = usuario.correo
            session["rol"] = usuario.rol

            if usuario.rol in ["admin", "vendedor"]:
                return redirect("/inventario")

            return redirect("/")

        error = "Correo o contraseña incorrectos."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# =========================
# USUARIOS
# =========================

@app.route("/usuarios", methods=["GET", "POST"])
def usuarios():
    if not session.get("rol"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return redirect("/inventario")

    mensaje = None
    error = None

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        correo = (request.form.get("correo") or "").strip().lower()
        password = request.form.get("password") or ""
        rol = request.form.get("rol") or "vendedor"

        if rol not in ["admin", "vendedor"]:
            rol = "vendedor"

        if not nombre or not correo or not password:
            error = "Todos los campos son obligatorios."

        elif Usuario.query.filter_by(correo=correo).first():
            error = "Ese correo ya está registrado."

        else:
            nuevo_usuario = Usuario(
                nombre=nombre,
                correo=correo,
                password_hash=generate_password_hash(password),
                rol=rol,
                activo=True
            )

            db.session.add(nuevo_usuario)
            db.session.commit()

            mensaje = "Usuario creado correctamente."

    usuarios_lista = Usuario.query.order_by(Usuario.id.desc()).all()

    return render_template(
        "usuarios.html",
        usuarios=usuarios_lista,
        mensaje=mensaje,
        error=error
    )


@app.route("/usuarios/desactivar/<int:id>")
def desactivar_usuario(id):
    if not session.get("rol"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return redirect("/inventario")

    usuario = Usuario.query.get_or_404(id)

    if usuario.correo != session.get("correo"):
        usuario.activo = False
        db.session.commit()

    return redirect("/usuarios")


@app.route("/usuarios/activar/<int:id>")
def activar_usuario(id):
    if not session.get("rol"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return redirect("/inventario")

    usuario = Usuario.query.get_or_404(id)
    usuario.activo = True
    db.session.commit()

    return redirect("/usuarios")


@app.route("/usuarios/eliminar/<int:id>")
def eliminar_usuario(id):
    if not session.get("rol"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return redirect("/inventario")

    usuario = Usuario.query.get_or_404(id)

    if usuario.correo != session.get("correo"):
        db.session.delete(usuario)
        db.session.commit()

    return redirect("/usuarios")


@app.route("/usuarios/reset/<int:id>", methods=["POST"])
def reset_password_usuario(id):
    if not session.get("rol"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return redirect("/inventario")

    usuario = Usuario.query.get_or_404(id)

    nueva_password = request.form.get("nueva_password") or ""

    if len(nueva_password) < 6:
        return redirect("/usuarios")

    usuario.password_hash = generate_password_hash(nueva_password)
    db.session.commit()

    return redirect("/usuarios")


# =========================
# ADMIN / AGREGAR PRODUCTO
# =========================

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    if request.method == "POST":
        archivos = request.files.getlist("fotos")
        nombre_archivo = guardar_varias_imagenes(archivos)

        if not nombre_archivo:
            archivo = request.files.get("imagen") or request.files.get("foto")
            nombre_archivo = guardar_imagen(archivo)

        nuevo = Producto(
            marca=request.form.get("marca"),
            modelo=request.form.get("modelo"),
            anio=int(request.form.get("anio") or 0),
            autoparte=request.form.get("autoparte") or request.form.get("nombre"),
            observaciones=request.form.get("observaciones"),
            transmision=request.form.get("transmision"),
            motor=request.form.get("motor"),
            propiedad=request.form.get("propiedad"),
            tipo=request.form.get("tipo"),
            lado=request.form.get("lado"),
            origen=request.form.get("origen"),
            costo_venta=float(request.form.get("costo_venta") or request.form.get("precio") or 0),
            stock=int(request.form.get("stock") or 1),
            foto=nombre_archivo,
            link_ml=request.form.get("link_ml"),
            estado="disponible"
        )

        db.session.add(nuevo)
        db.session.commit()

        return redirect("/inventario")

    productos = Producto.query.all()
    total = Producto.query.count()
    disponibles = Producto.query.filter(Producto.stock > 0).count()
    agotados = Producto.query.filter(Producto.stock == 0).count()

    return render_template(
        "admin.html",
        productos=productos,
        total=total,
        disponibles=disponibles,
        agotados=agotados
    )


# =========================
# INVENTARIO
# =========================

@app.route("/inventario")
def inventario():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    productos = Producto.query.all()

    return render_template(
        "inventario.html",
        productos=productos
    )


# =========================
# EDITAR
# =========================

@app.route("/editar/<int:id>", methods=["GET", "POST"])
def editar(id):
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    producto = Producto.query.get_or_404(id)

    if request.method == "POST":
        producto.marca = request.form.get("marca")
        producto.modelo = request.form.get("modelo")
        producto.anio = int(request.form.get("anio") or 0)
        producto.autoparte = request.form.get("autoparte") or request.form.get("nombre")
        producto.observaciones = request.form.get("observaciones")
        producto.transmision = request.form.get("transmision")
        producto.motor = request.form.get("motor")
        producto.propiedad = request.form.get("propiedad")
        producto.tipo = request.form.get("tipo")
        producto.lado = request.form.get("lado")
        producto.origen = request.form.get("origen")
        producto.costo_venta = float(
            request.form.get("costo_venta")
            or request.form.get("precio")
            or 0
        )
        producto.link_ml = request.form.get("link_ml")

        archivos = request.files.getlist("fotos")
        nuevas_fotos = guardar_varias_imagenes(archivos)

        if nuevas_fotos:
            producto.foto = nuevas_fotos
        else:
            archivo = request.files.get("foto")

            if archivo and archivo.filename:
                nombre_archivo = guardar_imagen(archivo)
                producto.foto = nombre_archivo

        db.session.commit()

        return redirect("/inventario")

    return render_template(
        "editar.html",
        producto=producto
    )


# =========================
# STOCK
# =========================

@app.route("/stock/<int:id>", methods=["GET", "POST"])
def ajustar_stock(id):
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    producto = Producto.query.get_or_404(id)

    if request.method == "POST":
        nuevo_stock = int(request.form.get("stock") or 0)

        producto.stock = nuevo_stock

        if producto.stock > 0:
            producto.estado = "disponible"
        else:
            producto.estado = "agotado"

        db.session.commit()

        return redirect("/inventario")

    return render_template(
        "stock.html",
        producto=producto
    )


@app.route("/sumar/<int:id>")
def sumar_stock(id):
    producto = Producto.query.get_or_404(id)
    producto.stock += 1
    producto.estado = "disponible"
    db.session.commit()
    return redirect("/inventario")


@app.route("/restar/<int:id>")
def restar_stock(id):
    producto = Producto.query.get_or_404(id)

    if producto.stock > 0:
        producto.stock -= 1

    if producto.stock == 0:
        producto.estado = "agotado"

    db.session.commit()
    return redirect("/inventario")


@app.route("/agotado/<int:id>")
def marcar_agotado(id):
    producto = Producto.query.get_or_404(id)
    producto.stock = 0
    producto.estado = "agotado"
    db.session.commit()
    return redirect("/inventario")


@app.route("/disponible/<int:id>")
def marcar_disponible(id):
    producto = Producto.query.get_or_404(id)
    producto.stock = 1
    producto.estado = "disponible"
    db.session.commit()
    return redirect("/inventario")


@app.route("/eliminar/<int:id>")
def eliminar(id):
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    producto = Producto.query.get_or_404(id)

    db.session.delete(producto)
    db.session.commit()

    return redirect("/inventario")


@app.route("/eliminar_masivo", methods=["POST"])
def eliminar_masivo():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    ids = request.form.getlist("productos")

    if ids:
        Producto.query.filter(
            Producto.id.in_(ids)
        ).delete(
            synchronize_session=False
        )

        db.session.commit()

    return redirect("/inventario")


# =========================
# DASHBOARD
# =========================

@app.route("/dashboard")
def dashboard():
    if session.get("rol") != "admin":
        return redirect("/inventario")

    productos = Producto.query.all()
    creditos_pendientes = CreditoVenta.query.filter(
        CreditoVenta.estado == "credito"
    ).all()

    total = Producto.query.count()
    disponibles = Producto.query.filter(Producto.stock > 0).count()
    agotados = Producto.query.filter(Producto.stock == 0).count()

    vendidas = Movimiento.query.filter(Movimiento.estado == "vendido").count()
    prestadas = Movimiento.query.filter(Movimiento.estado == "prestado").count()
    credito = len(creditos_pendientes)

    facturadas = Movimiento.query.filter(
        Movimiento.factura.isnot(None),
        Movimiento.factura != "",
        Movimiento.factura != "no"
    ).count()

    valor_total = 0
    resumen_propiedades = {}

    for producto in productos:
        precio = producto.costo_venta or 0
        stock = producto.stock or 0
        valor_producto = precio * stock
        valor_total += valor_producto

        propiedad = (producto.propiedad or "SIN PROPIEDAD").strip().upper()

        if not propiedad:
            propiedad = "SIN PROPIEDAD"

        if propiedad not in resumen_propiedades:
            resumen_propiedades[propiedad] = {
                "piezas": 0,
                "stock": 0,
                "valor": 0
            }

        resumen_propiedades[propiedad]["piezas"] += 1
        resumen_propiedades[propiedad]["stock"] += stock
        resumen_propiedades[propiedad]["valor"] += valor_producto

    resumen_propiedades = dict(
        sorted(
            resumen_propiedades.items(),
            key=lambda item: item[0]
        )
    )

    for propiedad in resumen_propiedades:
        resumen_propiedades[propiedad]["valor"] = round(
            resumen_propiedades[propiedad]["valor"],
            2
        )

    monto_ventas_contado = db.session.query(
        db.func.coalesce(db.func.sum(Movimiento.total), 0)
    ).filter(
        Movimiento.estado == "vendido"
    ).scalar()

    monto_abonos_credito = db.session.query(
        db.func.coalesce(db.func.sum(AbonoCreditoVenta.monto), 0)
    ).scalar()

    monto_vendido = float(monto_ventas_contado or 0) + float(monto_abonos_credito or 0)

    monto_credito = 0

    for credito_item in creditos_pendientes:
        abonado, saldo = calcular_saldo_credito(credito_item)
        monto_credito += saldo

    return render_template(
        "dashboard.html",
        total=total,
        disponibles=disponibles,
        agotados=agotados,
        valor_total=round(valor_total, 2),
        vendidas=vendidas,
        prestadas=prestadas,
        credito=credito,
        facturadas=facturadas,
        monto_vendido=round(float(monto_vendido or 0), 2),
        monto_credito=round(float(monto_credito or 0), 2),
        resumen_propiedades=resumen_propiedades
    )


# =========================
# VENTAS / CRÉDITOS / PRÉSTAMOS
# =========================

@app.route("/ventas", methods=["GET", "POST"])
def ventas():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    mensaje = None
    error = None

    if request.method == "POST":
        producto_id = request.form.get("producto_id")
        tipo_movimiento = request.form.get("tipo_movimiento")
        cantidad = int(request.form.get("cantidad") or 1)

        producto = Producto.query.get_or_404(producto_id)

        if cantidad <= 0:
            error = "La cantidad debe ser mayor a 0."

        elif cantidad > producto.stock:
            error = f"No hay suficiente stock. Stock actual: {producto.stock}"

        else:
            precio_unitario = producto.costo_venta or 0
            total_movimiento = calcular_total_movimiento(producto, cantidad)
            fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            producto.stock -= cantidad

            movimiento = Movimiento(
                producto_id=producto.id,
                tipo=tipo_movimiento,
                estado=tipo_movimiento,
                cliente=request.form.get("vendido_a"),
                cantidad=cantidad,
                precio_unitario=precio_unitario,
                total=total_movimiento,
                metodo_pago=request.form.get("metodo_pago"),
                factura=request.form.get("factura"),
                fecha_salida=request.form.get("fecha_salida"),
                fecha_prestamo=request.form.get("fecha_prestamo"),
                comentarios=request.form.get("comentarios_venta"),
                fecha_registro=fecha_registro
            )

            db.session.add(movimiento)
            db.session.flush()

            if tipo_movimiento == "credito":
                credito_nuevo = CreditoVenta(
                    movimiento_id=movimiento.id,
                    producto_id=producto.id,
                    cliente=request.form.get("vendido_a"),
                    cantidad=cantidad,
                    precio_unitario=precio_unitario,
                    total=total_movimiento,
                    estado="credito",
                    fecha_salida=request.form.get("fecha_salida"),
                    comentarios=request.form.get("comentarios_venta")
                )

                db.session.add(credito_nuevo)

            if producto.stock <= 0:
                producto.stock = 0
                producto.estado = tipo_movimiento
            else:
                producto.estado = "disponible"

            producto.vendido_a = request.form.get("vendido_a")
            producto.metodo_pago = request.form.get("metodo_pago")
            producto.factura = request.form.get("factura")
            producto.fecha_salida = request.form.get("fecha_salida")
            producto.fecha_prestamo = request.form.get("fecha_prestamo")

            producto.comentarios_venta = (
                f"Cantidad: {cantidad}. "
                f"{request.form.get('comentarios_venta') or ''}"
            )

            db.session.commit()

            mensaje = f"Movimiento registrado correctamente. Cantidad: {cantidad}. Total: ${total_movimiento}"

    productos = Producto.query.filter(Producto.stock > 0).all()

    movimientos = Movimiento.query.order_by(
        Movimiento.id.desc()
    ).all()

    return render_template(
        "ventas.html",
        productos=productos,
        movimientos=movimientos,
        mensaje=mensaje,
        error=error
    )


# =========================
# CRÉDITOS / ABONOS
# =========================

@app.route("/creditos")
def creditos():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    creditos_bd = CreditoVenta.query.order_by(
        CreditoVenta.id.desc()
    ).all()

    creditos_lista = []

    for credito_item in creditos_bd:
        abonado, saldo = calcular_saldo_credito(credito_item)

        if saldo <= 0 and credito_item.estado != "pagado":
            credito_item.estado = "pagado"

            if credito_item.movimiento:
                credito_item.movimiento.estado = "pagado"

            db.session.commit()

        creditos_lista.append({
            "id": credito_item.id,
            "producto": credito_item.producto,
            "cliente": credito_item.cliente,
            "cantidad": credito_item.cantidad,
            "total": round(credito_item.total or 0, 2),
            "abonado": round(abonado, 2),
            "saldo": round(saldo, 2),
            "estado": credito_item.estado,
            "fecha_salida": credito_item.fecha_salida,
            "comentarios": credito_item.comentarios
        })

    return render_template(
        "creditos.html",
        creditos=creditos_lista
    )


@app.route("/abonar/<int:id>", methods=["POST"])
def abonar_credito(id):
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    credito_item = CreditoVenta.query.get_or_404(id)

    if credito_item.estado == "pagado":
        return redirect("/creditos")

    try:
        monto = float(request.form.get("monto") or 0)
    except ValueError:
        monto = 0

    if monto <= 0:
        return redirect("/creditos")

    abonado_actual, saldo_actual = calcular_saldo_credito(credito_item)

    if saldo_actual <= 0:
        credito_item.estado = "pagado"

        if credito_item.movimiento:
            credito_item.movimiento.estado = "pagado"

        db.session.commit()
        return redirect("/creditos")

    if monto > saldo_actual:
        monto = saldo_actual

    abono = AbonoCreditoVenta(
        credito_id=credito_item.id,
        monto=monto,
        fecha=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        comentario=f"Abono registrado por {session.get('usuario', 'usuario')}"
    )

    db.session.add(abono)

    nuevo_abonado = abonado_actual + monto
    total_credito = credito_item.total or 0

    if nuevo_abonado >= total_credito:
        credito_item.estado = "pagado"

        if credito_item.movimiento:
            credito_item.movimiento.estado = "pagado"
            credito_item.movimiento.metodo_pago = "Crédito liquidado"

    db.session.commit()

    return redirect("/creditos")


@app.route("/marcar_pagado/<int:id>")
def marcar_pagado(id):
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    credito_item = CreditoVenta.query.get_or_404(id)

    if credito_item.estado == "pagado":
        return redirect("/creditos")

    abonado_actual, saldo = calcular_saldo_credito(credito_item)

    if saldo > 0:
        abono = AbonoCreditoVenta(
            credito_id=credito_item.id,
            monto=saldo,
            fecha=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            comentario=f"Liquidado por {session.get('usuario', 'usuario')}"
        )
        db.session.add(abono)

    credito_item.estado = "pagado"

    if credito_item.movimiento:
        credito_item.movimiento.estado = "pagado"
        credito_item.movimiento.metodo_pago = "Crédito liquidado"

    db.session.commit()

    return redirect("/creditos")


# =========================
# PRÉSTAMOS
# =========================

@app.route("/prestamos")
def prestamos():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    prestamos = Movimiento.query.filter(
        Movimiento.estado == "prestado"
    ).order_by(
        Movimiento.id.desc()
    ).all()

    return render_template(
        "prestamos.html",
        prestamos=prestamos
    )


@app.route("/devolver/<int:id>")
def devolver_prestamo(id):
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    movimiento = Movimiento.query.get_or_404(id)
    producto = movimiento.producto

    if movimiento.estado != "prestado":
        return redirect("/prestamos")

    movimiento.estado = "devuelto"

    if producto:
        producto.stock += movimiento.cantidad or 1

        if producto.stock > 0:
            producto.estado = "disponible"

    db.session.commit()

    return redirect("/prestamos")


# =========================
# EXCEL
# =========================

@app.route("/excel", methods=["GET", "POST"])
def excel():
    if not session.get("rol") in ["admin", "vendedor"]:
        return redirect("/login")

    mensaje = None

    if request.method == "POST":
        archivo = (
            request.files.get("archivo_excel")
            or request.files.get("excel")
        )

        archivo_zip = (
            request.files.get("imagenes_zip")
            or request.files.get("zip")
            or request.files.get("imagenes")
        )

        imagenes_disponibles = extraer_zip_imagenes(archivo_zip)

        if archivo and archivo.filename:
            workbook = load_workbook(archivo)
            hoja = workbook.active

            encabezados = []

            for celda in hoja[1]:
                encabezado = str(celda.value or "").strip().lower()
                encabezados.append(encabezado)

            total_cargados = 0

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                datos = dict(zip(encabezados, fila))

                if not datos.get("autoparte") and not datos.get("marca"):
                    continue

                fotos_excel = []

                for campo in ["foto", "foto1", "foto2", "foto3", "foto4", "foto5"]:
                    nombre_foto = datos.get(campo)

                    if nombre_foto:
                        nombre_foto = secure_filename(str(nombre_foto).strip())

                        if nombre_foto:
                            if not imagenes_disponibles or nombre_foto in imagenes_disponibles:
                                fotos_excel.append(nombre_foto)

                    if len(fotos_excel) == 5:
                        break

                nuevo = Producto(
                    marca=datos.get("marca"),
                    modelo=datos.get("modelo"),
                    anio=int(datos.get("año") or datos.get("anio") or 0),
                    observaciones=datos.get("observaciones"),
                    transmision=datos.get("transmision"),
                    motor=datos.get("motor"),
                    autoparte=datos.get("autoparte"),
                    propiedad=datos.get("propiedad"),
                    tipo=datos.get("tipo"),
                    lado=datos.get("lado"),
                    origen=datos.get("origen"),
                    costo_venta=float(datos.get("costo de venta") or datos.get("costo_venta") or 0),
                    foto="|".join(fotos_excel),
                    link_ml=datos.get("link mercado libre") or datos.get("link_ml"),
                    stock=int(datos.get("stock") or 1),
                    estado="disponible",
                    comentarios_venta=datos.get("seccion de las ventas y comentarios"),
                    metodo_pago=datos.get("metodo de pago"),
                    factura=datos.get("factura"),
                    fecha_salida=str(datos.get("fecha de la salida") or ""),
                    fecha_prestamo=str(datos.get("fecha de prestamos") or "")
                )

                db.session.add(nuevo)
                total_cargados += 1

            db.session.commit()

            mensaje = f"Se cargaron {total_cargados} productos correctamente."

    return render_template("excel.html", mensaje=mensaje)


# =========================
# CREAR TABLAS Y ADMIN INICIAL
# =========================

with app.app_context():
    db.create_all()

    if Usuario.query.count() == 0:
        admin_inicial = Usuario(
            nombre="Administrador",
            correo=os.getenv("ADMIN_CORREO", "admin@autopartesch.com"),
            password_hash=generate_password_hash(os.getenv("ADMIN_PASSWORD", "Admin12345")),
            rol="admin",
            activo=True
        )

        db.session.add(admin_inicial)
        db.session.commit()

        print("====================================")
        print("ADMIN INICIAL CREADO")
        print("Correo:", os.getenv("ADMIN_CORREO", "admin@autopartesch.com"))
        print("Contraseña temporal:", os.getenv("ADMIN_PASSWORD", "Admin12345"))
        print("CAMBIA ESTOS DATOS EN PRODUCCIÓN")
        print("====================================")


# =========================
# RUN
# =========================

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port)